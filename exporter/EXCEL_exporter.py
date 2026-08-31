# /exporter/EXCEL_exporter.py

import logging
import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .base import CourseExporter

logger = logging.getLogger(__name__)

WEEKDAY_MAP = {
    '星期一': 1, '星期二': 2, '星期三': 3, '星期四': 4,
    '星期五': 5, '星期六': 6, '星期日': 7,
    '周一': 1, '周二': 2, '周三': 3, '周四': 4,
    '周五': 5, '周六': 6, '周日': 7,
}
WEEKDAY_CN = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

_CN_NUM = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '日': 7}

_P_HTML_RE = re.compile(r'<p>(.*?)</p>', re.DOTALL)
_WEEKDAY_RE = re.compile(r'(?:星期|周)([一二三四五六日])')
_SECTION_RE = re.compile(r'第(\d+)-(\d+)节')
_WEEKS_RE = re.compile(r'(\d+-\d+)\s*(?:单|双)?周')
_LOCATION_RE = re.compile(r'第\d+-\d+节\s*(.*?)\s*$')


class ExcelExporter(CourseExporter):

    def __init__(self, courses, file_path, max_section: int = 12):
        super().__init__(courses, file_path)
        self.max_section = max_section

    # ------------------------------------------------------------------
    # 解析
    # ------------------------------------------------------------------

    def _parse_time(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        解析课程时间，返回 {课程名: [时间条目]}。

        时间条目字段：
            weeks     周次字符串，如 '1-16'
            week_type 'all' | 'odd' | 'even'（每周 / 单周 / 双周）
            weekday   1-7（星期一 = 1）
            start     起始节次（从 1 开始）
            end       结束节次
            location  上课地点，如 '一教107'
        """
        result: Dict[str, List[Dict[str, Any]]] = {}

        for course in self.courses:
            name = course.get('kcmc', '')
            if not name:
                continue

            entries = []
            for content in _P_HTML_RE.findall(course.get('pkjgmx', '')):
                content = content.strip()
                if not content:
                    continue

                weekday_m = _WEEKDAY_RE.search(content)
                section_m = _SECTION_RE.search(content)
                weeks_m = _WEEKS_RE.search(content)
                if not (weekday_m and section_m and weeks_m):
                    continue

                if '单周' in content:
                    week_type = 'odd'
                elif '双周' in content:
                    week_type = 'even'
                else:
                    week_type = 'all'

                location_m = _LOCATION_RE.search(content)
                location = location_m.group(1).strip() if location_m else ''

                entries.append({
                    'weeks': weeks_m.group(1),
                    'week_type': week_type,
                    'weekday': _CN_NUM[weekday_m.group(1)],
                    'start': int(section_m.group(1)),
                    'end': int(section_m.group(2)),
                    'location': location,
                })

            if entries:
                result[name] = entries

        self.date_data = result
        return result

    # ------------------------------------------------------------------
    # 导出
    # ------------------------------------------------------------------

    def export(self) -> None:
        """导出课程表到 Excel 文件。"""
        date_data = self._parse_time()

        # 动态计算最大节次（至少覆盖数据，且不低于默认值）
        max_section = self.max_section
        for entries in date_data.values():
            for e in entries:
                max_section = max(max_section, e['end'])

        wb = Workbook()
        ws = wb.active
        ws.title = '课程表'

        # --- 样式常量 ---
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', start_color='4F81BD', end_color='4F81BD')
        body_font = Font(name='微软雅黑', size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='999999')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        weekend_fill = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')

        # --- 表头：节次 + 星期 ---
        ws.cell(row=1, column=1, value='节次')
        for col, day in enumerate(WEEKDAY_CN, start=2):
            ws.cell(row=1, column=col, value=day)

        # --- 节次列 ---
        for sec in range(1, max_section + 1):
            ws.cell(row=sec + 1, column=1, value=f'第{sec}节')

        # --- 填充课程（先写值，再合并，避免对 MergedCell 赋值） ---
        # grid[(weekday, section)] -> 该格的所有课程文本
        grid: Dict[tuple, List[str]] = {}
        for name, entries in date_data.items():
            for e in entries:
                text = self._course_cell_text(name, e)
                for s in range(e['start'], e['end'] + 1):
                    grid.setdefault((e['weekday'], s), []).append(text)

        for name, entries in date_data.items():
            for e in entries:
                text = self._course_cell_text(name, e)
                col = e['weekday'] + 1
                for s in range(e['start'], e['end'] + 1):
                    cell = ws.cell(row=s + 1, column=col)
                    cell.value = f"{cell.value}\n{text}" if cell.value else text

        # --- 合并：同一课程在同一星期内连续且独占的节次 ---
        for name, entries in date_data.items():
            for e in entries:
                text = self._course_cell_text(name, e)
                col = e['weekday'] + 1
                r0, r1 = e['start'] + 1, e['end'] + 1
                if r1 > r0 and all(
                    grid[(e['weekday'], s)] == [text]
                    for s in range(e['start'], e['end'] + 1)
                ):
                    ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)

        # --- 样式 ---
        for row in ws.iter_rows(min_row=1, max_row=max_section + 1, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = center
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = body_font
                    if cell.column >= 7:  # 周六、周日
                        cell.fill = weekend_fill

        # --- 列宽 / 行高 ---
        ws.column_dimensions['A'].width = 8
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.row_dimensions[1].height = 24
        for r in range(2, max_section + 2):
            ws.row_dimensions[r].height = 42

        # 冻结首行首列
        ws.freeze_panes = 'B2'

        # --- 保存 ---
        file_path = self.file_path
        if not str(file_path).lower().endswith('.xlsx'):
            file_path = str(file_path) + '.xlsx'
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        logger.info(f'课程表已导出：{file_path}')

    # ------------------------------------------------------------------
    # 工具
    # ------------------------------------------------------------------

    @staticmethod
    def _course_cell_text(name: str, entry: Dict[str, Any]) -> str:
        """生成单元格文本：课程名 / 周次（单双周）/ 教室"""
        lines = [name]
        week = f"{entry['weeks']}周"
        if entry['week_type'] == 'odd':
            week += '（单周）'
        elif entry['week_type'] == 'even':
            week += '（双周）'
        lines.append(week)
        if entry.get('location'):
            lines.append(entry['location'])
        return '\n'.join(lines)
